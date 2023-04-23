                    Работа с таблицами EXEL
================ Модуль openpyxl ======================================================

pip install --user openpyxl


Открытие документов:
--------------------    
>>> import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx') # Функция для открытия excel документов
>>> type(wb)
<class 'openpyxl.workbook.workbook.Workbook'> # Возвращает значение типа Workbook

Для работы с excel файлом, необходимо чтобы он находился в текущем каталоге
Определение папки явл. текущим каталогом:

>>> import os
>>> print(os.getcwd())
# Можно воспользоваться os.chdir() для смены рабочего каталога


Получение списка листов рабочей книги:
--------------------------------------
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames                               # Имена листов книги
['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3']                        # Конкретный лист
sheet
<Worksheet "Sheet3">
type(sheet)
<class 'openpyxl.worksheet.worksheet.Worksheet'> # Каждый лист представлен обьектом Worksheet
sheet.title
'Sheet3'
anotherSheet = wb.active                     # Активный лист
anotherSheet
<Worksheet "Sheet1">


Получение ячеек рабочих листов:
--------------------------------------
Терминал должен запускаться из той же папки где находится файл excel

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']                            # Получение листа книги
sheet['A1']                                     # Получение ячейки листа
<Cell 'Sheet1'.A1>
sheet['A1'].value                               # Получение значения ячейки
datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']                                 # Получение другой ячейки
c.value
'Apples'

# Получение строки, столбца и значения из ячейки
'Строка %s, Столбец %s : %s' % (c.row, c.column, c.value)
'Строка 1, Столбец 2 : Apples'
'Ячейка %s : %s' % (c.coordinate, c.value)
'Ячейка B1 : Apples'
sheet['C1'].value
73



sheet.cell(row=1, column=2)
<Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value
'Apples'
for i in range(1, 8, 2): # Проходим каждую вторую строку
    print(i, sheet.cell(row=i, column=2).value)

    
1 Apples
3 Pears
5 Apples
7 Strawberries

# Размер листа
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
sheet.max_row         # Наибольший номер строки
7
sheet.max_column      # Наибольший номер столбца
3



Преобразование буквенных и числовых обозначений столбцов:
---------------------------------------------------------
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1)          # Translate column 1 to a letter.
'A'
get_column_letter(2)
'B'
get_column_letter(27)
'AA'
get_column_letter(900)
'AHP'
wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
'C'
column_index_from_string('A') # Get A's number.
1
column_index_from_string('AA')
27


Получение строк и столбцов рабочих листов:
------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
tuple(sheet['A1':'C3'])  # Все ячейки от A1 до C3
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- КОНЕЦ СТРОКИ ---')

    
A1 2015-04-05 13:34:02
B1 Apples
C1 73
--- КОНЕЦ СТРОКИ ---
A2 2015-04-05 03:41:23
B2 Cherries
C2 85
--- КОНЕЦ СТРОКИ ---
A3 2015-04-06 12:46:51
B3 Pears
C3 14
--- КОНЕЦ СТРОКИ ---

# Перебираем весь столбик
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
list(sheet.columns)[1] # Ячейки второго столбца
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

    
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries


================= Reading Data from a Spreadsheet =====================================

#! python3
# readCensusExcel.py - формирует таблицу с данными о численности
# населения и количестве переписных районов в каждом округе

import openpyxl, pprint
print('Открытие рабочей книги...')
wb = openpyxl.load_workbook('censuspopdata.xlsx') # Открываем файл
sheet = wb['Population by Census Tract']          # Получаем лист с данными переписи
countyData = {}



print('Чтение строк...')
for row in range(2, sheet.max_row + 1):
    # В каждой строке электронной таблицы содержатся данные для одного переписного района
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Гарантируем наличие ключа для данного штата
    countyData.setdefault(state, {})
    # Гарантируем наличие ключа для данного округа штата
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Каждая строка представляет один переписной район, поэтому увеличиваем результат на единицу
    countyData[state][county]['tracts'] += 1
    # Увеличение численности населения округа на количество жителей переписного района
    countyData[state][county]['pop'] += int(pop)

# Открытие нового текстового файла и запись в него содержимого словаря countyData
print('Запись результатов...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Готово.')

By outputting it to a text file named census2010.py
you’ve generated a Python program from your Python program

>>> import census2010
>>> census2010.allData['AK']['Anchorage']
{'pop': 291826, 'tracts': 55}
>>> anchoragePop = census2010.allData['AK']['Anchorage']['pop']
>>> print('Население округа Анкоридж в 2010 году - ' + str(anchoragePop))  
Население округа Анкоридж в 2010 году - 291826



==================== Создание и сохранение документов Excel ======================

import openpyxl
wb = openpyxl.Workbook() # Создаём пустую рабочую книгу
wb.sheetnames            # Она содержит один лист
['Sheet']
sheet = wb.active
sheet.title
'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet'  # Меняем название листа
wb.sheetnames
['Spam Bacon Eggs Sheet']


Предпологается что файл example.xlsx находится в текущем каталоге
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam, Spam, Spam'
wb.save('example_copy.xlsx')    # Сохраняем новую рабочую книгу
 


==================== Сохранение и удаление рабочих листов ======================











